# stock_monitor/report.py

from datetime import date
import pandas as pd
from supabase import create_client
from stock_monitor.config import SUPABASE_URL, SUPABASE_KEY, EMAIL_SENDER, EMAIL_RECIPIENT, SENDGRID_API_KEY, STORE_NAMES

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

def get_snapshots(snapshot_type):
    today = date.today().isoformat()
    response = (
        supabase.table("stock_snapshots")
        .select("*")
        .eq("snapshot_date", today)
        .eq("snapshot_type", snapshot_type)
        .execute()
    )
    return response.data


def get_thresholds():
    response = supabase.table("prescription_thresholds").select("*").execute()
    return {
        (row["sphere"], row["cylinder"]): row["threshold"]
        for row in response.data
    }


def build_report():
    morning = get_snapshots("morning")
    eod = get_snapshots("eod")
    thresholds = get_thresholds()

    morning_df = pd.DataFrame(morning)
    eod_df = pd.DataFrame(eod)

    merged = morning_df.merge(
        eod_df,
        on=["store_id", "sphere", "cylinder"],
        suffixes=("_morning", "_eod")
    )

    merged["consumed"] = merged["quantity_morning"] - merged["quantity_eod"]
    merged["threshold"] = merged.apply(
        lambda row: thresholds.get((row["sphere"], row["cylinder"]), None),
        axis=1
    )
    merged["below_threshold"] = merged["quantity_eod"] < merged["threshold"]
    merged["store_id"] = merged["store_id"].map(STORE_NAMES) 

    return merged[[
        "store_id",
        "sphere",
        "cylinder",
        "quantity_morning",
        "quantity_eod",
        "consumed",
        "threshold",
        "below_threshold"
    ]]


def save_excel(df):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    from io import BytesIO

    df = df.copy()
    df["below_threshold"] = df["below_threshold"].map({True: "Yes", False: "No"})

    alerts_df = df[df["below_threshold"] == "Yes"]

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        alerts_df.to_excel(writer, sheet_name="Alerts", index=False)
        df.to_excel(writer, sheet_name="Full Report", index=False)

    buffer.seek(0)

    wb = load_workbook(buffer)
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    ws = wb["Full Report"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[-1].value == "Yes":
            for cell in row:
                cell.fill = red_fill

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def send_email(buffer):
    from sendgrid import SendGridAPIClient
    from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition
    import base64

    encoded = base64.b64encode(buffer.read()).decode()
    filename = f"stock_report_{date.today().isoformat()}.xlsx"

    attachment = Attachment(
        FileContent(encoded),
        FileName(filename),
        FileType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        Disposition("attachment")
    )

    message = Mail(
        from_email=EMAIL_SENDER,
        to_emails=EMAIL_RECIPIENT,
        subject=f"Lens Stock Report - {date.today().isoformat()}",
        plain_text_content="Please find attached today's lens stock report for the SIOC mobiles."
    )
    message.attachment = attachment

    sg = SendGridAPIClient(SENDGRID_API_KEY)
    response = sg.send(message)
    print(f"Email sent, status code: {response.status_code}")


def run_report():
    df = build_report()
    buffer = save_excel(df)
    send_email(buffer)